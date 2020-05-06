#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
import pandas as pd
import pathlib
from sys import platform
from os import listdir
from os.path import isfile, join
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows as pd_to_xl

path = str(pathlib.Path().absolute())
prefix = '/'
if platform == 'darwin':
    path = '/Users/belindalu/Documents/EsobiTimeSheet'
if platform == 'win32':
    prefix = '\\'
path = path + prefix
mode_code = ['1', '2', '3']


def execute():
    user_in = None
    while user_in != '3':
        user_in = switch_mode()
        if user_in == '1':
            file_path = get_path()
            mode_a(file_path)
        if user_in == '2':
            file_path = get_path()
            mode_b(file_path)
        if user_in not in mode_code:
            input('代號錯誤, enter 返回 MENU')
            continue
        if user_in != '3':
            input('輸入 enter 後返回 MENU')


def get_path():
    files = [f for f in listdir(path) if isfile(join(path, f))]
    file_dict = {}
    print('請輸入time_sheet檔對應的id：')
    i = 0
    for f in files:
        file_dict[i] = f
        print(f'{i}) {f}')
        i = i + 1
    file_id = int(input('檔案id:').strip())
    if file_id not in file_dict:
        raise ValueError
    return path + file_dict[file_id]


def switch_mode():
    print('選擇功能或離開:')
    print('1) 查詢工時未填滿人員')
    print('2) mode_b')
    print('3) QUIT')
    user_in = input('請輸入功能代號：')
    return str(user_in)


def mode_a(file_path):
    # 1. 取得檔案內容
    time_sheet = pd.read_excel(file_path)
    export_sheet = pd.DataFrame()
    export_sheet['member'], export_sheet['work_time'] = time_sheet['登記人'], time_sheet['耗時']
    export_sheet = export_sheet.groupby(['member']).sum()

    # 2. 取得member_path
    member_path = path + 'members.xlsx'

    # 3. 篩選出今日上班人員
    members = pd.read_excel(member_path)
    members = members[members.check == 'Y'].drop('check', axis=1)

    # 4. 排除重複輸入的人員
    members = members.drop_duplicates(keep='first')

    # 5. join 人員清單與工時
    final_data = pd.merge(members, export_sheet, how='left', on='member')
    final_data.fillna(0, inplace=True)
    final_data = final_data[final_data.work_time < 8]
    final_data['loss_time'] = 8 - final_data['work_time']
    print(final_data)


def mode_b(file_path):
    # 1. 取得檔案內容
    time_sheet = pd.read_excel(file_path)
    time_sheet.sort_values(by=['登記人'], inplace=True)
    export_sheet = pd.DataFrame()
    wb = Workbook()
    ws = wb.active
    for r in pd_to_xl(time_sheet, index=True, header=True):
        ws.append(r)
    #
    # rows = (
    #     (88, 46, 57),
    #     (88, 38, 12),
    #     (88, 59, 78),
    #     (88, 21, 98),
    #     (88, 18, 43),
    #     (88, 15, 67)
    # )
    # for row in rows:
    #     ws.append(row)
    # df = pd.DataFrame(ws.values)
    # df
    # ws.merge_cells('C3:C29')
    # df2 = pd.DataFrame(ws.values)
    # df2

    # rows = (
    #     (88, 46, 57),
    #     (88, 38, 12),
    #     (88, 59, 78),
    #     (88, 21, 98),
    #     (88, 18, 43),
    #     (88, 15, 67)
    # )
    # for row in rows:
    #     ws.append(row)
    # df = pd.DataFrame(ws.values)
    # df
    # ws.merge_cells('A1:A6')
    # df2 = pd.DataFrame(ws.values)
    # df2

    # ws.unmerge_cells('A2:D2')
    # # or equivalently
    # ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
    # ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)

    # Save the file
    wb.save("xxxxx.xlsx")

    # def to_merge_cell():
    pass


try:
    execute()
except:
    print('你他媽的別在那邊亂')
finally:
    print('******** BYE BYE ********')
